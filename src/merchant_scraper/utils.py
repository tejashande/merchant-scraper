"""
Utility functions for the merchant scraper.
"""

import logging
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.settings import OUTPUT_DIR

from .models import Merchant


def setup_logging():
    """Configure logging for the application."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    )
    return logging.getLogger(__name__)


def save_to_excel(merchants: List[Merchant], filename: str) -> None:
    """
    Save merchant data to Excel file with formatted headers and table.

    Args:
        merchants: List of Merchant objects
        filename: Name of the output file (without extension)
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Merchants"

    # Define header names
    headers = {
        "name": "Business Name",
        "address": "Address",
        "latitude": "Latitude",
        "longitude": "Longitude",
        "business_types": "Business Types",
        "mcc_code": "MCC Code",
        "mcc_category": "MCC Category",
        "place_id": "Place ID",
        "rating": "Rating",
        "user_ratings_total": "Total Ratings",
        "price_level": "Price Level",
        "is_open": "Open Now",
        "phone": "Phone",
        "website": "Website",
    }

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_num, (field, header) in enumerate(headers.items(), 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

        # Auto-adjust column width
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = max(
            len(header) + 2,  # Add some padding
            ws.column_dimensions[column_letter].width or 0,
        )

    # Write data
    for row_num, merchant in enumerate(merchants, 2):
        data = {
            "name": merchant.name,
            "address": merchant.address,
            "latitude": merchant.latitude,
            "longitude": merchant.longitude,
            "business_types": ", ".join(merchant.business_types),
            "mcc_code": merchant.mcc_code,
            "mcc_category": merchant.mcc_category,
            "place_id": merchant.place_id,
            "rating": merchant.rating,
            "user_ratings_total": merchant.user_ratings_total,
            "price_level": merchant.price_level,
            "is_open": "Yes" if merchant.is_open else "No",
            "phone": merchant.phone or "",
            "website": merchant.website or "",
        }

        for col_num, (field, value) in enumerate(data.items(), 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Auto-adjust column width for data
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = max(
                len(str(value)) + 2,  # Add some padding
                ws.column_dimensions[column_letter].width or 0,
            )

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Save the file
    filepath = f"{OUTPUT_DIR}/{filename}.xlsx"
    wb.save(filepath)
