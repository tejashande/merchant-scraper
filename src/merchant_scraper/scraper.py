"""
Main scraper module for fetching merchant data from Google Places API.
"""

import os
import time
from datetime import datetime
from typing import Dict, List, Optional, Set, Tuple

import googlemaps

from config.settings import (
    DEFAULT_RADIUS,
    GOOGLE_PLACES_API_KEY,
    MAX_REQUESTS_PER_DAY,
    MIN_DELAY_BETWEEN_REQUESTS,
    OUTPUT_DIR,
    PAGINATION_DELAY,
)

from .constants import B2C_BUSINESS_TYPES
from .models import Merchant
from .utils import setup_logging

logger = setup_logging()


def get_mcc_from_google_place_type(place_type: str) -> Tuple[str, str]:
    """
    Get MCC code and category for a Google Places business type.

    Args:
        place_type: Google Places business type

    Returns:
        Tuple of (MCC code, category)
    """
    from src.mcc_codes import get_mcc_from_google_place_type as _get_mcc

    return _get_mcc(place_type)


class MerchantScraper:
    """Class to scrape merchant data from Google Places API."""

    def __init__(
        self, api_key: str = GOOGLE_PLACES_API_KEY, output_dir: str = OUTPUT_DIR
    ):
        """
        Initialize the scraper.

        Args:
            api_key: Google Places API key
            output_dir: Directory to save output files
        """
        self.gmaps = googlemaps.Client(key=api_key)
        self.output_dir = output_dir
        self.seen_place_ids: Set[str] = set()
        self.request_count = 0
        self.last_request_time = 0

        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)

        logger.info("MerchantScraper initialized")

    def _rate_limit(self):
        """Implement rate limiting for API requests."""
        current_time = time.time()
        time_since_last_request = current_time - self.last_request_time

        if time_since_last_request < MIN_DELAY_BETWEEN_REQUESTS:
            sleep_time = MIN_DELAY_BETWEEN_REQUESTS - time_since_last_request
            time.sleep(sleep_time)

        self.last_request_time = time.time()
        self.request_count += 1

        if self.request_count >= MAX_REQUESTS_PER_DAY:
            logger.error("Daily API quota exceeded")
            raise Exception("Daily API quota exceeded")

    def _make_api_request(self, func, *args, **kwargs):
        """Make an API request with rate limiting."""
        self._rate_limit()
        return func(*args, **kwargs)

    def _is_b2c_business(self, business_types: List[str]) -> bool:
        """
        Check if a business is B2C based on its types.

        Args:
            business_types: List of business types from Google Places API

        Returns:
            True if the business is B2C, False otherwise
        """
        # Flatten all B2C business types into a set for quick lookup
        b2c_types = set()
        for category in B2C_BUSINESS_TYPES.values():
            b2c_types.update(category)

        # Check if any of the business types are in our B2C set
        return any(business_type in b2c_types for business_type in business_types)

    def _map_business_type_to_mcc(self, business_types: List[str]) -> Tuple[str, str]:
        """
        Map Google Places business types to MCC codes.

        Args:
            business_types: List of business types from Google Places API

        Returns:
            Tuple of (MCC code, category)
        """
        # Try to find a matching MCC code for each business type
        for business_type in business_types:
            try:
                code, category = get_mcc_from_google_place_type(business_type)
                return code, category
            except ValueError:
                continue

        # If no match found, use default MCC code
        logger.warning(f"No MCC code found for business types: {business_types}")
        return "5399", "Miscellaneous General Merchandise"

    def _process_place_data(self, place: Dict) -> Optional[Merchant]:
        """
        Process raw place data into merchant format.

        Args:
            place: Raw place data from Google Places API

        Returns:
            Merchant object or None if processing fails
        """
        try:
            # Get business types from place data
            business_types = place.get("types", [])

            # Skip if not a B2C business
            if not self._is_b2c_business(business_types):
                return None

            # Skip if we've already seen this place
            place_id = place.get("place_id", "")
            if place_id in self.seen_place_ids:
                return None
            self.seen_place_ids.add(place_id)

            # Get additional details
            details = self._make_api_request(
                self.gmaps.place, place_id, fields=["formatted_phone_number", "website"]
            )

            # Map business types to MCC codes
            mcc_code, mcc_category = self._map_business_type_to_mcc(business_types)

            return Merchant(
                name=place.get("name", ""),
                address=place.get("vicinity", ""),
                latitude=place.get("geometry", {}).get("location", {}).get("lat", 0),
                longitude=place.get("geometry", {}).get("location", {}).get("lng", 0),
                business_types=business_types,
                mcc_code=mcc_code,
                mcc_category=mcc_category,
                place_id=place_id,
                rating=place.get("rating", 0),
                user_ratings_total=place.get("user_ratings_total", 0),
                price_level=place.get("price_level", 0),
                is_open=place.get("opening_hours", {}).get("open_now", False),
                phone=details["result"].get("formatted_phone_number"),
                website=details["result"].get("website"),
            )
        except Exception as e:
            logger.error(f"Error processing place data: {str(e)}")
            return None

    def fetch_merchants(
        self, location: str, radius: int = DEFAULT_RADIUS
    ) -> List[Merchant]:
        """
        Fetch merchant data from Google Places API.

        Args:
            location: Location to search (address or coordinates)
            radius: Search radius in meters

        Returns:
            List of Merchant objects
        """
        try:
            # Geocode the location
            geocode_result = self._make_api_request(self.gmaps.geocode, location)
            if not geocode_result:
                raise ValueError(f"Could not geocode location: {location}")

            location_lat = geocode_result[0]["geometry"]["location"]["lat"]
            location_lng = geocode_result[0]["geometry"]["location"]["lng"]

            # Search for places in each B2C category
            merchants = []
            for category, business_types in B2C_BUSINESS_TYPES.items():
                logger.info(f"Searching for {category} businesses...")

                for business_type in business_types:
                    try:
                        logger.info(
                            f"Searching for {business_type} business type in {category} category..."
                        )

                        # Search for places
                        places_result = self._make_api_request(
                            self.gmaps.places_nearby,
                            location=(location_lat, location_lng),
                            radius=radius,
                            type=business_type,
                        )

                        # Process results
                        for place in places_result.get("results", []):
                            merchant = self._process_place_data(place)
                            if merchant:
                                merchants.append(merchant)

                        # Handle pagination with delay
                        while "next_page_token" in places_result:
                            time.sleep(
                                PAGINATION_DELAY
                            )  # Wait before next page request

                            try:
                                places_result = self._make_api_request(
                                    self.gmaps.places_nearby,
                                    location=(location_lat, location_lng),
                                    radius=radius,
                                    type=business_type,
                                    page_token=places_result["next_page_token"],
                                )

                                for place in places_result.get("results", []):
                                    merchant = self._process_place_data(place)
                                    if merchant:
                                        merchants.append(merchant)
                            except Exception as e:
                                logger.warning(
                                    f"Error in pagination for {business_type}: {str(e)}"
                                )
                                break  # Break pagination loop on error

                    except Exception as e:
                        logger.error(f"Error processing {business_type}: {str(e)}")
                        continue

            logger.info(f"Found {len(merchants)} B2C merchants")
            return merchants
        except Exception as e:
            logger.error(f"Error fetching merchants: {str(e)}")
            return []

    def run(self, location: str, radius: int = DEFAULT_RADIUS) -> None:
        """
        Run the scraper and save results to Excel.

        Args:
            location: Location to search (address or coordinates)
            radius: Search radius in meters
        """
        try:
            # Fetch merchants
            merchants = self.fetch_merchants(location, radius)
            if not merchants:
                logger.warning(f"No merchants found for location: {location}")
                return

            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"merchants_{timestamp}"

            # Save to Excel
            self._save_to_excel(merchants, filename)

        except Exception as e:
            logger.error(f"Error running scraper: {str(e)}")

    def _save_to_excel(self, merchants: List[Merchant], filename: str) -> None:
        """
        Save merchant data to an Excel file.

        Args:
            merchants: List of Merchant objects
            filename: Name of the output file (without extension)
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active

        # Define headers
        headers = {
            "name": "Name",
            "address": "Address",
            "latitude": "Latitude",
            "longitude": "Longitude",
            "business_types": "Business Types",
            "mcc_code": "MCC Code",
            "mcc_category": "MCC Category",
            "place_id": "Place ID",
            "rating": "Rating",
            "user_ratings_total": "Total Ratings",
            "price_level": "Price Level",
            "is_open": "Is Open",
            "phone": "Phone",
            "website": "Website",
        }

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_num, (field, header) in enumerate(headers.items(), 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

            # Auto-adjust column width
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = max(
                len(header) + 2,  # Add some padding
                ws.column_dimensions[column_letter].width or 0,
            )

        # Write data
        for row_num, merchant in enumerate(merchants, 2):
            data = {
                "name": merchant.name,
                "address": merchant.address,
                "latitude": merchant.latitude,
                "longitude": merchant.longitude,
                "business_types": ", ".join(merchant.business_types),
                "mcc_code": merchant.mcc_code,
                "mcc_category": merchant.mcc_category,
                "place_id": merchant.place_id,
                "rating": merchant.rating,
                "user_ratings_total": merchant.user_ratings_total,
                "price_level": merchant.price_level,
                "is_open": "Yes" if merchant.is_open else "No",
                "phone": merchant.phone or "",
                "website": merchant.website or "",
            }

            for col_num, (field, value) in enumerate(data.items(), 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Save the file
        output_path = os.path.join(self.output_dir, f"{filename}.xlsx")
        wb.save(output_path)
        logger.info(f"Saved merchant data to {output_path}")
